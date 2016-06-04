from openpyxl import load_workbook

wb=load_workbook("BaseMedicamentos.xlsx")
ws=wb.active

for row in ws.iter_rows('B2:B9'):
    for cell in row:
        print cell 
